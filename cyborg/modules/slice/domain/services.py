import datetime
import json
import logging
import os
import shutil
import time
from typing import List, Optional, Tuple

import numpy as np
import openpyxl
from PIL import Image
from xlsxwriter import Workbook
from xlsxwriter.format import Format

from cyborg.app.request_context import request_context
from cyborg.app.settings import Settings
from cyborg.consts.common import Consts
from cyborg.consts.tct import TCTConsts
from cyborg.infra.cache import cache
from cyborg.infra.fs import fs
from cyborg.infra.session import transaction
from cyborg.libs.heimdall.dispatch import open_slide
from cyborg.modules.slice.application.tasks import update_clarity
from cyborg.modules.slice.domain.entities import CaseRecordEntity, SliceEntity, ReportConfigEntity
from cyborg.modules.slice.domain.repositories import CaseRecordRepository, ReportConfigRepository
from cyborg.modules.slice.domain.value_objects import SliceStartedStatus
from cyborg.seedwork.domain.value_objects import AIType
from cyborg.utils.image import rotate_jpeg
from cyborg.utils.strings import camel_to_snake

logger = logging.getLogger(__name__)


class SliceDomainService(object):

    def __init__(
            self, repository: CaseRecordRepository, report_config_repository: ReportConfigRepository):
        super(SliceDomainService, self).__init__()
        self.repository = repository
        self.report_config_repository = report_config_repository

    def recognize_label_text(self, label_path: str, slice_label_path: Optional[str], label_rec_mode: int):
        if os.path.exists(label_path) and label_rec_mode is not None and label_rec_mode > 1:
            try:
                label_img = Image.open(label_path)
                if label_img.mode == "RGBA":
                    label_img.convert('RGB')
                    label_img.save(label_path)
                from cyborg.libs.label_ocr.label_rec import label_recognition
                return label_recognition(label_path, slice_label_path, mode=label_rec_mode)
            except Exception as e:
                logger.error('label ocr error :' + str(e))
                return ''
        return ''

    def update_clarity(self, slice_id: int, slice_file_path: str):
        try:
            from cyborg.modules.slice.utils.clarify import blur_check
            logger.info('计算清晰度: %s' % slice_file_path)
            slide = open_slide(slice_file_path)
            clarity_score = blur_check(slide)
        except Exception as e:
            logger.exception(e)
            logger.info(f'清晰度计算失败: {e}')
            clarity_score = 0

        slice_entity = self.repository.get_slice_by_id(slice_id)
        if slice_entity:
            slice_entity.update_data(clarity=clarity_score)
            if self.repository.save_slice(slice_entity):
                logger.info('清晰度计算完成: %s: %s' % (slice_file_path, clarity_score))

    def create_slice(
            self, case_id: str, file_id: str, company_id: str, slide_path: str, file_name: str, local_file_name: str,
            upload_type: str, tool_type: str, label_rec_mode: int, user_file_path: str, cover_slice_number: bool,
            operator: str, high_through: bool = False
    ) -> Optional[dict]:

        _ext = os.path.splitext(local_file_name)[-1].lower()  # 上传切片文件后缀
        if _ext not in Consts.ALLOWED_EXTENSIONS:
            return

        slide_save_name = os.path.join(slide_path, local_file_name)
        if _ext == ".jpeg":
            rotate_jpeg(slide_save_name)
        slide = open_slide(slide_save_name)

        try:
            if _ext == ".svs":
                objective_rate = slide.slide.properties['aperio.AppMag'] + 'X'
            else:
                raise Exception('cannot read AppMag from slide')
        except Exception as e:
            logger.warning(e)
            if slide.mpp:
                if _ext not in Consts.OBJECTIVE_RATE_DICT:
                    _ext = 'other'
                objective_rate = Consts.OBJECTIVE_RATE_DICT[_ext][
                    round(float(slide.mpp) / 0.242042)]
            else:
                objective_rate = ''

        label_path = os.path.join(slide_path, "label.png")
        slide.save_label(label_path)  # 先关闭掉 后边再调
        slice_label_path = os.path.join(slide_path, 'slice_label.jpg')

        label_text = self.recognize_label_text(label_path, slice_label_path, label_rec_mode)

        clarity_score = '-'

        slide_size = os.path.getsize(slide_save_name)

        try:
            thumbnail_image = slide.get_thumbnail(Settings.THUMBNAIL_BOUNDING)
            thumbnail_image.save(os.path.join(slide_path, "thumbnail.jpeg"))
        except Exception:
            logger.warning(f'获取缩略图失败: {case_id}-{file_id}')
            pass

        slice_entity = SliceEntity(raw_data={
            'fileid': file_id,
            "caseid": case_id,
            'filename': file_name,
            'name': file_name,
            'loaded': slide_size,
            'total': slide_size,
            'mppx': slide.mpp,
            'mppy': slide.mpp,
            'height': slide.height,
            'width': slide.width,
            'objective_rate': objective_rate,
            "company": company_id,
            "type": "slice",
            "update_time": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "operator": operator,
            "slice_number": label_text,
            "is_has_label": 1 if os.path.exists(slice_label_path) and os.path.getsize(
                slice_label_path) else 0,
            "clarity": clarity_score,
            "is_cs": 1 if upload_type == 'cs' else 0,
            "high_through": high_through,
            "tool_type": tool_type,
            "user_file_path": user_file_path,
            "user_file_folder": user_file_path.split('\\')[-2] if user_file_path else ''
        })
        self.repository.save_slice(slice_entity)

        update_clarity(slice_entity.id, slide_save_name)

        if high_through:
            record = self.repository.get_record_by_case_id(case_id=case_id, company=company_id)
            if not record:  # 高通量模式, 若未创建病例则要创建病例文件夹并且在数据库中插入病例信息
                record = CaseRecordEntity(raw_data=dict(
                    caseid=case_id,
                    company=company_id,
                    sample_num=case_id if cover_slice_number else (label_text or case_id),
                    stage=0,
                    slice_count=0,
                    started=0,
                    state=1,
                    report=2,
                    create_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
                    update_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
                ))
            else:
                record.update_data(update_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"))

            record.update_data(slice_count=record.slice_count + 1)
            self.repository.save(record)
        return {
            "width": slide.width,
            "height": slide.height,
            "mppx": slide.mpp,
            "mppy": slide.mpp,
            "ai_dict": {},
            "state": 1,
            "total": slide_size,
            "loaded": slide_size,
            "objective_rate": objective_rate,
            "slice_number": label_text,
            "clarity": clarity_score,
            "userFilePath": user_file_path
        }

    @transaction
    def update_slice_info(
            self, case_id: str, file_id: str, high_through: bool, info: dict) -> bool:
        slice = self.repository.get_slice(
            case_id=case_id, file_id=file_id, company=request_context.current_company)
        record = self.repository.get_record_by_case_id(
            case_id=case_id, company=request_context.current_company)
        if high_through:
            if 'slice_number' in info:
                record.update_data(slice_number=info['slice_number'])
            slice.update_data(high_through=1)

        info['update_time'] = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

        if slice:
            cover = False
            if 'started' in info:
                del info['started']
            if 'ai_suggest' in info:
                del info['ai_suggest']
            if 'radius' in info:
                del info['radius']
            if 'is_solid' in info:
                del info['is_solid']
            if 'clarity' in info:
                del info['clarity']

            if slice.alg == 'her2' and not info.get('check_result'):
                cover = True

            for k, v in info.items():
                k = camel_to_snake(k)
                if k == 'id':
                    slice.update_data(fileid=v)

                elif k == 'ai_suggest':
                    if info.get('check_result') and slice.update_ai_result(v):
                        cover = True
                    slice.update_data(ai_suggest=json.dumps(v))

                elif v is not None:
                    if isinstance(v, str):
                        slice.update_data(**{k: v.strip()})
                    else:
                        slice.update_data(**{k: v})

            if cover:
                slice.update_data(check_result='')
            self.repository.save_slice(slice)

        else:  # 新建切片信息
            slice = SliceEntity(**info)
            if high_through:
                slice.update_data(high_through=1)
            self.repository.save_slice(slice)

            if record:
                record.update_data(update_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"))
                record.update_data(slice_num=record.slice_num + 1)
                self.repository.save(record)
        return True

    def update_mark_config(
            self, case_id: str, file_id: str, radius: Optional[float] = None, is_solid: Optional[int] = None
    ) -> bool:
        entity = self.repository.get_slice(
            case_id=case_id, file_id=file_id, company=request_context.current_company)
        if not entity:
            return False
        if radius is not None:
            entity.update_data(radius=radius)
        if is_solid is not None:
            entity.update_data(is_solid=is_solid)

        return self.repository.save_slice(entity)

    def reset_ai_status(
            self, case_id: str, file_id: str, company_id: str) -> Optional[SliceEntity]:
        entity = self.repository.get_slice(case_id=case_id, file_id=file_id, company=company_id)
        if not entity:
            return None
        entity.update_data(
            started=0,
            ai_suggest='',
            ai_status=0
        )
        if self.repository.save_slice(entity):
            return entity
        return None

    def update_ai_status(
            self, case_id: str, file_id: str, company_id: str, status: SliceStartedStatus,
            ai_name: Optional[str] = None, upload_batch_number: Optional[str] = None,
            template_id: Optional[int] = None, ip_address: Optional[str] = None
    ) -> Optional[SliceEntity]:
        entity = self.repository.get_slice(case_id=case_id, file_id=file_id, company=company_id)
        if not entity:
            return None

        entity.update_data(
            started=status,
            ai_suggest='',
            ai_status=0
        )
        if ai_name is not None:
            ai_dict = entity.ai_dict
            ai_dict[ai_name + 'Started'] = True
            entity.update_data(ai_dict=ai_dict, alg=ai_name, tool_type=ai_name)
        if upload_batch_number is not None:
            entity.update_data(upload_batch_number=upload_batch_number)
        if template_id is not None:
            entity.update_data(template_id=template_id)
        if ip_address is not None:
            entity.update_data(ipaddress=ip_address)

        entity.update_data(update_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"))

        if self.repository.save_slice(entity):
            return entity
        return None

    def get_slice(
            self, case_id: str, file_id: str, company_id: str) -> Tuple[int, str, Optional[SliceEntity]]:
        slice = self.repository.get_slice(case_id=case_id, file_id=file_id, company=company_id)
        if not slice:
            return 2, 'no such file or directory', None

        return 0, '', slice

    @cache.cache('cyborg:slice_analysis:slice_file_info:{case_id}:{file_id}', expire=cache.ONE_HOUR)
    def get_slice_file_info(self, case_id: str, file_id: str, company_id: str) -> Optional[dict]:
        slice = self.repository.get_slice(
            case_id=case_id, file_id=file_id, company=company_id)
        if slice:
            slide = open_slide(slice.slice_file_path)
            return slide.to_dict()
        return None

    def remove_slice_files(self, slice: SliceEntity) -> float:
        total_size = fs.get_dir_size(slice.record_dir)
        if fs.path_exists(slice.slice_dir):
            shutil.rmtree(slice.slice_dir, ignore_errors=True)
        elif fs.path_exists(slice.attachment_dir):
            shutil.rmtree(slice.attachment_dir, ignore_errors=True)
        return total_size - fs.get_dir_size(slice.record_dir)

    def remove_ai_image_files(self, slice: SliceEntity):
        dna_index_img_path = fs.path_join(slice.slice_file_path, 'dna_index.png')
        dna_scatter_img_path = fs.path_join(slice.slice_file_path, 'scatterplot.png')
        if fs.path_exists(dna_index_img_path):
            shutil.rmtree(dna_index_img_path, ignore_errors=True)
        elif fs.path_exists(dna_scatter_img_path):
            shutil.rmtree(dna_scatter_img_path, ignore_errors=True)

    @transaction
    def delete_slice(self, slice: SliceEntity) -> float:
        self.repository.delete_slice(file_id=slice.fileid, company_id=slice.company)
        record = self.repository.get_record_by_case_id(case_id=slice.caseid, company=slice.company)
        if record and record.slice_count is not None:
            record.slice_count -= 1
            self.repository.save(record)

        return self.remove_slice_files(slice)

    def update_template_id(self, case_id: str, file_id: str, template_id: int) -> Tuple[int, str]:
        err_code, message, entity = self.get_slice(
            case_id=case_id, file_id=file_id, company_id=request_context.current_company)

        if entity:
            entity.update_data(template_id=template_id)
            if not self.repository.save_slice(entity):
                err_code, message = 1, 'update slice failed'

        return err_code, message

    def write_records_to_excel(
            self, records: List[CaseRecordEntity], headers: List[str], workbook: Workbook) -> Optional[str]:

        worksheet = workbook.add_worksheet('sheet1')
        common_format = workbook.add_format(
            {'bold': True,
             'border': 1,
             'align': 'center',
             'valign': 'vcenter',
             'text_wrap': True}
        )

        positive_format = workbook.add_format(
            {'bold': True,
             'border': 1,
             'align': 'center',
             'valign': 'vcenter',
             'font_color': 'red',
             'text_wrap': True}
        )

        negative_format = workbook.add_format(
            {'bold': True,
             'border': 1,
             'align': 'center',
             'valign': 'vcenter',
             'font_color': 'green',
             'text_wrap': True}
        )

        def auto_cell_format(cell_text) -> Format:
            return positive_format if cell_text and '阳性' in cell_text else negative_format

        # 设定列标题，调整宽度
        for col, title in enumerate(headers):
            worksheet.write(0, col, title, common_format)
            if '最后更新' in headers:
                worksheet.set_column(headers.index('最后更新'), headers.index('最后更新'), 25)
            if '创建时间' in headers:
                worksheet.set_column(headers.index('创建时间'), headers.index('创建时间'), 25)
            if '切片标签' in headers:
                worksheet.set_column(headers.index('切片标签'), headers.index('切片标签'), 20)

        row_items = [(record, slice) for record in records for slice in record.slices]
        for idx, (record, slice) in enumerate(row_items):
            row = idx + 1
            for col, title in enumerate(headers):
                cell_format = common_format
                write_label = ''
                if title == '样本号':
                    write_label = record.caseid
                elif title == '姓名':
                    write_label = record.name
                elif title == '性别':
                    write_label = record.gender
                elif title == '年龄':
                    write_label = str(record.age)
                elif title == '取样部位':
                    write_label = record.sample_part
                elif title == '样本类型':
                    write_label = record.sample_type
                elif title == '切片数量':
                    write_label = '1'
                elif title == '状态':
                    write_label = {'0': '未处理', '1': '处理中', '2': '已处理', '3': '处理异常'}.get(str(slice.ai_status), '')
                elif title == '切片标签':
                    slice_label_path = slice.slice_label_path
                    if os.path.exists(slice_label_path) and os.path.getsize(slice_label_path):
                        # worksheet.insert_image(row, col, slice_label_path, {'x_scale': 0.11, 'y_scale': 0.1})
                        pass
                elif title == '切片编号':
                    write_label = slice.slice_number
                elif title == '文件名':
                    write_label = slice.filename
                elif title == '切片文件夹':
                    write_label = slice.user_file_folder
                elif title == 'AI模块':
                    write_label = Consts.ALGOR_DICT.get(slice.alg) if slice.alg in Consts.ALGOR_DICT else slice.alg
                elif title == 'AI建议':
                    write_label = slice.ai_suggest
                    cell_format = auto_cell_format(write_label)
                elif title == '复核结果':
                    write_label = slice.check_result.strip().strip(';')
                    cell_format = auto_cell_format(write_label)
                elif title == '最后更新':
                    update_time = record.update_time
                    if update_time:
                        try:
                            utc_date = datetime.datetime.strptime(update_time, '%Y-%m-%dT%H:%M:%SZ')
                            local_date = utc_date + datetime.timedelta(hours=8)
                            local_date_str = local_date.strftime('%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            utc_time = datetime.datetime.strptime(update_time, '%Y-%m-%dT%H:%M:%S.%fZ')
                            local_date_str = utc_time + datetime.timedelta(hours=8)
                            local_date_str = local_date_str.strftime('%Y-%m-%d %H:%M:%S')
                        write_label = local_date_str
                elif title == '创建时间':
                    create_time = record.create_time
                    write_label = ''
                    if create_time:
                        try:
                            utc_date = datetime.datetime.strptime(create_time, '%Y-%m-%dT%H:%M:%SZ')
                            local_date = utc_date + datetime.timedelta(hours=8)
                            local_date_str = local_date.strftime('%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            utc_time = datetime.datetime.strptime(create_time, '%Y-%m-%dT%H:%M:%S.%fZ')
                            local_date_str = utc_time + datetime.timedelta(hours=8)
                            local_date_str = local_date_str.strftime('%Y-%m-%d %H:%M:%S')
                        write_label = local_date_str
                elif title == '操作人':
                    write_label = slice.operator
                elif title == '最终结果':
                    write_label = slice.check_result.strip().strip(';') or slice.ai_suggest
                    cell_format = auto_cell_format(write_label)

                worksheet.write(row, col, write_label, cell_format)

        return None

    def validate_slice_file(self, slide_path: str, file_name: str) -> bool:
        try:
            if os.path.splitext(file_name)[-1] in ['.mdsx', '.svs']:  # 上传麦克奥迪和海德星切片的处理
                current_dir = os.path.join(slide_path, os.path.splitext(file_name)[0])
                if os.path.exists(current_dir) and os.path.isdir(current_dir):
                    for file in os.listdir(current_dir):
                        if file.endswith('.mdsx') or file.endswith('.svs'):
                            link_file = os.path.join(slide_path, file_name)
                            if os.path.exists(link_file):
                                os.remove(link_file)
                            os.symlink(
                                os.path.join(os.path.join(current_dir, file)), link_file)
                            break
            open_slide(os.path.join(slide_path, file_name))
            return True
        except Exception as e:
            logger.warning(f'切片名为{file_name}文件损坏: {e}')
            return False

    def save_record(
            self,
            user_name: str,
            company_id: str,
            case_id: Optional[str] = None,
            report_info: Optional[str] = None,
            attachments: Optional[List[dict]] = None,
            record_name: Optional[str] = None,
            age: Optional[str] = None,
            gender: Optional[str] = None,
            cancer_type: Optional[str] = None,
            family_history: Optional[str] = None,
            medical_history: Optional[str] = None,
            generally_seen: Optional[str] = None,
            sample_num: Optional[str] = None,
            sample_part: Optional[str] = None,
            sample_time: Optional[str] = None,
            sample_collect_date: Optional[str] = None,
            sample_type: Optional[str] = None,
            inspection_hospital: Optional[str] = None,
            inspection_doctor: Optional[str] = None,
            opinion: Optional[str] = None,
            stage: int = 0, started: int = 0, state: int = 1, report: str = '2',
            **_
    ) -> Optional[CaseRecordEntity]:
        record = self.repository.get_record_by_case_id(case_id=case_id, company=company_id) if case_id else None
        if not record:
            case_id = case_id or CaseRecordEntity.gen_case_id()
            record = CaseRecordEntity(raw_data=dict(
                caseid=case_id,
                company=company_id,
                create_time=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
            ))

        for attachment in attachments or []:
            fileid = attachment.get('id', None)
            name = attachment.get('name', None)
            filename = attachment.get('filename', None)

            slice_entity = self.repository.get_slice(case_id=case_id, file_id=fileid, company=company_id)
            if not slice_entity:
                slice_entity = SliceEntity()

            loaded = attachment.get('loaded', 0)
            total = attachment.get('total', 0)
            stain = attachment.get('stain', None)
            state = attachment.get('state', 1)
            ajax_token = json.dumps(attachment.get('ajaxToken', {}))
            path = attachment.get('path', '')
            res_obj = {
                'caseid': case_id,
                'filename': filename,
                'name': name,
                'loaded': loaded,
                'total': total,
                'stain': stain,
                'state': state,
                'mppx': 0.0,
                'mppy': 0.0,
                'height': 0,
                'width': 0,
                'tool_type': None,
                'started': 0,
                'objective_rate': '',
                'radius': 1,  # 默认系数是1
                'is_solid': 1,  # 默认标注模块中显示实心
                'fileid': fileid,
                'company': company_id,
                'ajax_token': ajax_token,
                'path': path,
                'type': "attachment",
                'update_time': datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
                'operator': user_name,
            }
            slice_entity.update_data(**res_obj)
            self.repository.save_slice(slice_entity)

        slices = self.repository.get_slices_by_case_id(case_id=case_id, company=company_id)
        slice_count = self.repository.get_slice_count_by_case_id(case_id=case_id, company=company_id)
        if sample_num is None:
            sample_num = slices[0].slice_number if slices else ''

        record.update_data(**{
            'caseid': case_id,
            'name': record_name,
            'age': age,
            'gender': gender,
            'cancer_type': cancer_type,
            'family_history': family_history,
            'medical_history': medical_history,
            'sample_num': sample_num,
            'sample_part': sample_part,
            'sample_time': sample_time,
            'sample_collect_date': sample_collect_date,
            'sample_type': sample_type,
            'generally_seen': generally_seen,
            'inspection_hospital': inspection_hospital,
            'inspection_doctor': inspection_doctor,
            'report_info': report_info,
            'opinion': opinion,
            'update_time': datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            'stage': stage,
            'slice_count': slice_count,
            'started': started,
            'state': state,
            'report': report
        })
        if self.repository.save(record):
            record.slices = slices
            return record
        return None

    def remove_record_files(self, record: CaseRecordEntity) -> float:
        total_size = fs.get_dir_size(record.data_dir)
        if os.path.exists(record.data_dir):
            shutil.rmtree(record.data_dir, ignore_errors=True)
        return total_size - fs.get_dir_size(record.data_dir)

    @transaction
    def delete_records(self, case_ids: List[str], company_id: str) -> float:
        freed_size = 0
        for case_id in case_ids:
            record = self.repository.get_record_by_case_id(case_id=case_id, company=company_id)
            if record:
                if self.repository.delete_record(case_id=case_id, company_id=company_id):
                    freed_size += self.remove_record_files(record)
        return freed_size

    def import_records(self, company_id: str) -> dict:
        import_cache_path = os.path.join(request_context.current_user.data_dir, 'importCache.xlsx')
        # 读取上传文件
        wb = openpyxl.load_workbook(import_cache_path)
        ws = wb.worksheets[0]  # 默认导入sheet1中的数据

        succeed_count = 0  # 新增成功计数
        failure_count = 0  # 插入失败计数（多个病例存在相同文件名跳过处理）
        skip_count = 0  # 跳过处理计数（找不到同名切片）
        for i in range(3, ws.max_row + 1):  # 从第三行开始读取sheet中的数据
            line_data = ws['a' + str(i): 'j' + str(i)]
            is_empty = False
            for j in line_data[0]:
                if j.value and not is_empty:
                    is_empty = False
            if not is_empty and line_data[0][7].value is not None:  # 过滤掉空行和文件名为空的行
                sample_num = line_data[0][0].value  # 样本号
                name = line_data[0][1].value  # 姓名
                gender = line_data[0][2].value  # 性别
                age = line_data[0][3].value  # 年龄
                sample_part = line_data[0][4].value  # 取样部位
                sample_type = line_data[0][5].value  # 样本类型
                slice_number = line_data[0][6].value  # 切片编号
                filename = line_data[0][7].value  # 文件名
                ai_suggest = line_data[0][8].value  # AI建议
                check_result = line_data[0][9].value  # 复核结果

                slices = self.repository.get_slices(file_name=filename, company=company_id)
                if not slices:
                    skip_count += 1
                elif len(slices) > 1:
                    failure_count += 1
                else:
                    slice = slices[0]
                    record = self.repository.get_record_by_case_id(slice.caseid, company=company_id)
                    record.update_data(
                        sample_num=sample_num,
                        name=name,
                        gender=gender,
                        age=age,
                        sample_part=sample_part,
                        sample_type=sample_type
                    )
                    self.repository.save(record)

                    slice.update_data(
                        slice_number=slice_number,
                        filename=filename,
                        ai_suggest=ai_suggest,
                        check_result=check_result
                    )
                    self.repository.save_slice(slice)
                    succeed_count += 1

        return {
            'succeed_count': succeed_count,
            'failure_count': failure_count,
            'skip_count': skip_count
        }

    def apply_ai_threshold(self, company: str, ai_type: AIType, threshold_range: int, threshold_value: float) -> bool:
        for prob, slice in self.repository.get_prob_list(company=company, ai_type=ai_type):
            ai_suggest = slice.ai_suggest.split(" ")
            if len(ai_suggest) == 1:
                pre_tbs_label, pre_else = '', ['']
            elif len(ai_suggest) == 2:
                pre_tbs_label, pre_else = ai_suggest[1].strip(), ['']
            elif len(ai_suggest) > 2:
                pre_tbs_label, pre_else = ai_suggest[1].strip(), ai_suggest[2:]
            else:
                continue

            prob_np = np.array(prob.to_list())
            pred_tbs = np.argmax(prob_np)
            tbs_label = TCTConsts.tct_multi_wsi_cls_dict_reverse[pred_tbs]
            if threshold_range == 0 and ('LSIL' in tbs_label or 'HSIL' in tbs_label or 'AGC' in tbs_label):
                continue
            else:
                diagnosis = '阳性' if 1 - prob_np[0] > threshold_value else '阴性'
                if tbs_label == 'NILM' and diagnosis == '阳性':
                    tbs_label = 'ASC-US'
                tbs_label = '' if diagnosis == '阴性' else tbs_label

                if '不满意' in pre_tbs_label:
                    tbs_label = tbs_label + '-样本不满意'
                pre_else = [pre_else] if not isinstance(pre_else, list) else pre_else
                new_ai_suggest = ' '.join([diagnosis, tbs_label]) + ' ' + ' '.join(pre_else)
                new_ai_suggest = new_ai_suggest.strip()
                slice.update_data(ai_suggest=new_ai_suggest)
                self.repository.save_slice(slice)
        return True

    @transaction
    async def free_up_space(self, end_time: str, company: str) -> bool:
        records = self.repository.get_records(end_time=end_time, company=request_context.current_company)
        for record in records:
            self.repository.delete_record(record.caseid, company_id=company)

        slices = self.repository.get_slices(case_ids=[record.caseid for record in records])
        for slice_entity in slices:
            self.repository.delete_slice(slice_entity.fileid, company_id=company)

        for record in records:
            await record.remove_data_dir()

        return True

    def get_report_config(self, company: str) -> Optional[ReportConfigEntity]:
        config = self.report_config_repository.get_by_company(company=company)
        if not config:
            config = ReportConfigEntity.new_entity(company=company)
            if not self.report_config_repository.save(config):
                return None

        return config


if __name__ == '__main__':
    path = '/Users/zhaoyu/dipath_data/company1/data/2023_05_19_16_05_11_102179/slices/8271014/鼻息肉.jpg'
    slide = open_slide(path)
    from cyborg.modules.slice.utils.clarify import blur_check
    blur_check(slide)
